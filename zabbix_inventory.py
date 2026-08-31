#!/usr/bin/env python3
"""
zabbix_inventory.py
====================

Generates an Excel (.xlsx) inventory of a Zabbix 7.0.4 installation using
only the Zabbix JSON-RPC API (no Ansible, no third-party Zabbix SDK).

Covers, per host: status, interfaces, proxy, host groups, templates, tags
and description. LLD/Discovery Rules, native host inventory, items and
triggers are intentionally out of scope for this first version -- see
README.md for the extension points.

Usage:
    python zabbix_inventory.py
    python zabbix_inventory.py --output inventario.xlsx
    python zabbix_inventory.py --url https://zabbix.example.com

Configuration is read from a .env file (see .env.example); CLI arguments
override .env values.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import sys
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

LOG = logging.getLogger("zabbix_inventory")

# ---------------------------------------------------------------------------
# Zabbix 7.0 API value mappings
#
# These are documented, stable enumerations of the Zabbix 7.0 API (verified
# against the official 7.0 API reference for host.get / proxy.get):
#   - Interface type:  1=Zabbix Agent, 2=SNMP, 3=IPMI, 4=JMX
#   - Host status:     0=Enabled, 1=Disabled
#   - monitored_by:    0=Zabbix server (Direct), 1=Proxy, 2=Proxy group
# ---------------------------------------------------------------------------

INTERFACE_TYPES = {
    "1": "Zabbix Agent",
    "2": "SNMP",
    "3": "IPMI",
    "4": "JMX",
}

HOST_STATUS = {
    "0": "Enabled",
    "1": "Disabled",
}

MONITORED_BY_SERVER = "0"
MONITORED_BY_PROXY = "1"
MONITORED_BY_PROXY_GROUP = "2"

DEFAULT_TIMEOUT = 30  # seconds
DEFAULT_OUTPUT = "inventario.xlsx"

HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)
WRAP_ALIGN_TOP = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN = Alignment(vertical="top")

MAX_COL_WIDTH = 60
MIN_COL_WIDTH = 10


class ConfigError(Exception):
    """Raised for configuration / credential problems (fatal, non-API)."""


class ZabbixAPIError(Exception):
    """Raised for any transport or JSON-RPC level error from the Zabbix API."""


# ---------------------------------------------------------------------------
# 1. Configuration
# ---------------------------------------------------------------------------

@dataclass
class Config:
    url: str
    token: Optional[str]
    user: Optional[str]
    password: Optional[str]
    output: str
    verify_ssl: bool = True


def _parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate an Excel inventory of a Zabbix 7.0.4 installation via the JSON-RPC API."
    )
    parser.add_argument("--url", help="Zabbix frontend base URL or full api_jsonrpc.php URL. Overrides ZABBIX_URL.")
    parser.add_argument("--output", help=f"Output .xlsx path (default: {DEFAULT_OUTPUT}).")
    return parser.parse_args(argv)


def _normalize_api_url(url: str) -> str:
    url = url.rstrip("/")
    if not url.endswith("api_jsonrpc.php"):
        url = f"{url}/api_jsonrpc.php"
    return url


def load_config(argv: Optional[List[str]] = None) -> Config:
    """Load configuration from .env, then apply CLI overrides (CLI wins)."""
    load_dotenv()
    args = _parse_args(argv)

    raw_url = args.url or os.getenv("ZABBIX_URL")
    if not raw_url:
        raise ConfigError("No Zabbix URL configured. Set ZABBIX_URL in .env or pass --url.")
    url = _normalize_api_url(raw_url)

    token = os.getenv("ZABBIX_TOKEN", "").strip() or None
    user = os.getenv("ZABBIX_USER", "").strip() or None
    password = os.getenv("ZABBIX_PASSWORD", "").strip() or None

    if not token and not (user and password):
        raise ConfigError(
            "No valid authentication method configured. Set ZABBIX_TOKEN, "
            "or both ZABBIX_USER and ZABBIX_PASSWORD, in your .env file."
        )

    verify_raw = os.getenv("ZABBIX_VERIFY_SSL", "true").strip().lower()
    verify_ssl = verify_raw not in ("false", "0", "no")
    if not verify_ssl:
        LOG.warning(
            "TLS certificate verification is DISABLED (ZABBIX_VERIFY_SSL=false). "
            "Only use this for trusted internal/self-signed environments -- traffic "
            "is still encrypted, but the server identity is not verified."
        )

    output = args.output or os.getenv("ZABBIX_OUTPUT") or DEFAULT_OUTPUT

    return Config(url=url, token=token, user=user, password=password, output=output, verify_ssl=verify_ssl)


# ---------------------------------------------------------------------------
# 2. Communication with Zabbix (pure requests, JSON-RPC 2.0)
# ---------------------------------------------------------------------------

class ZabbixClient:
    """Minimal JSON-RPC 2.0 client for the Zabbix API."""

    def __init__(self, url: str, verify_ssl: bool = True, timeout: int = DEFAULT_TIMEOUT):
        self.url = url
        self.timeout = timeout
        self.session = requests.Session()
        self.session.verify = verify_ssl
        self._auth_token: Optional[str] = None
        self._request_id = 0
        self.version: Optional[str] = None

    def _call(self, method: str, params: Optional[dict] = None, authenticated: bool = True) -> Any:
        self._request_id += 1
        payload = {
            "jsonrpc": "2.0",
            "method": method,
            "params": params or {},
            "id": self._request_id,
        }
        headers = {"Content-Type": "application/json-rpc"}
        if authenticated and self._auth_token:
            # Zabbix >= 6.4 accepts (and for tokens, requires) the Bearer
            # header instead of the legacy top-level "auth" field.
            headers["Authorization"] = f"Bearer {self._auth_token}"

        try:
            response = self.session.post(
                self.url, data=json.dumps(payload), headers=headers, timeout=self.timeout
            )
        except requests.exceptions.SSLError as exc:
            raise ZabbixAPIError(
                f"TLS certificate validation failed for {self.url}: {exc}. "
                "If this is an internal/self-signed server, set ZABBIX_VERIFY_SSL=false "
                "in .env (see README for the security implications)."
            ) from exc
        except requests.exceptions.ConnectionError as exc:
            raise ZabbixAPIError(f"Could not connect to {self.url}: {exc}") from exc
        except requests.exceptions.Timeout as exc:
            raise ZabbixAPIError(f"Timed out contacting {self.url} after {self.timeout}s: {exc}") from exc
        except requests.exceptions.RequestException as exc:
            raise ZabbixAPIError(f"Unexpected network error contacting {self.url}: {exc}") from exc

        if response.status_code != 200:
            raise ZabbixAPIError(
                f"Zabbix API returned HTTP {response.status_code} for method '{method}': "
                f"{response.text[:300]!r}"
            )

        try:
            data = response.json()
        except ValueError as exc:
            raise ZabbixAPIError(
                f"Zabbix API returned a non-JSON response for method '{method}' "
                f"(is the URL correct? got: {response.text[:200]!r})"
            ) from exc

        if "error" in data:
            err = data["error"]
            raise ZabbixAPIError(
                f"Zabbix API error on '{method}': {err.get('message', '')} - {err.get('data', '')}"
            )

        return data.get("result")

    def check_version(self) -> str:
        """apiinfo.version requires no authentication; used as a first connectivity check."""
        self.version = self._call("apiinfo.version", {}, authenticated=False)
        return self.version

    def login_with_token(self, token: str) -> None:
        self._auth_token = token

    def login_with_credentials(self, user: str, password: str) -> None:
        result = self._call(
            "user.login", {"username": user, "password": password}, authenticated=False
        )
        if not result:
            raise ZabbixAPIError("user.login did not return a session token.")
        self._auth_token = result

    def verify_authentication(self) -> None:
        """Cheapest authenticated call available: fetch a single hostid."""
        self._call("host.get", {"output": ["hostid"], "limit": 1})

    def logout(self) -> None:
        if self._auth_token:
            try:
                self._call("user.logout", {})
            except ZabbixAPIError:
                # Best-effort cleanup; not worth failing the run over.
                pass
            self._auth_token = None

    def host_get(self, params: dict) -> List[dict]:
        return self._call("host.get", params) or []

    def proxy_get(self, params: dict) -> List[dict]:
        return self._call("proxy.get", params) or []


def create_zabbix_client(config: Config) -> ZabbixClient:
    LOG.info("Connecting to Zabbix...")
    client = ZabbixClient(config.url, verify_ssl=config.verify_ssl)

    try:
        version = client.check_version()
    except ZabbixAPIError as exc:
        raise ZabbixAPIError(f"Could not reach the Zabbix API at {config.url}: {exc}") from exc
    LOG.info("Connected to Zabbix %s", version)

    if config.token:
        LOG.info("Authenticating with API token...")
        client.login_with_token(config.token)
    else:
        LOG.info("Authenticating with username/password...")
        try:
            client.login_with_credentials(config.user, config.password)
        except ZabbixAPIError as exc:
            raise ZabbixAPIError(f"Authentication failed: {exc}") from exc

    try:
        client.verify_authentication()
    except ZabbixAPIError as exc:
        raise ZabbixAPIError(f"Authentication check failed (invalid token or credentials?): {exc}") from exc

    return client


# ---------------------------------------------------------------------------
# 3. Retrieval
# ---------------------------------------------------------------------------

def get_hosts(client: ZabbixClient) -> List[dict]:
    """
    Retrieve ALL hosts (enabled and disabled) with every related object
    needed for the inventory, in a single API call.

    Uses selectHostGroups (selectGroups is deprecated since Zabbix 6.2),
    selectParentTemplates (directly linked templates -- see README for the
    note on template inheritance), selectInterfaces and selectTags.
    selectInventory is intentionally NOT used: native host inventory is out
    of scope for this first version.
    """
    LOG.info("Retrieving hosts...")
    hosts = client.host_get(
        {
            "output": [
                "hostid",
                "host",
                "name",
                "status",
                "proxyid",
                "monitored_by",
                "proxy_groupid",
                "description",
            ],
            "selectInterfaces": ["interfaceid", "type", "ip", "dns", "port", "main"],
            "selectHostGroups": ["groupid", "name"],
            "selectParentTemplates": ["templateid", "name"],
            "selectTags": ["tag", "value"],
        }
    )
    LOG.info("Retrieved %d hosts", len(hosts))
    return hosts


def get_proxy_lookup(client: ZabbixClient) -> Dict[str, str]:
    """One proxy.get call to resolve proxyid -> proxy name for all hosts at once."""
    LOG.info("Retrieving proxies...")
    proxies = client.proxy_get({"output": ["proxyid", "name"]})
    lookup = {p["proxyid"]: p.get("name", "") for p in proxies}
    LOG.info("Retrieved %d proxies", len(lookup))
    return lookup


# ---------------------------------------------------------------------------
# 4. Processing
# ---------------------------------------------------------------------------

@dataclass
class HostRow:
    hostid: str
    host: str
    name: str
    status: str
    proxy_display: str
    proxy_id: str
    groups: List[str] = field(default_factory=list)
    templates: List[str] = field(default_factory=list)
    tags: List[str] = field(default_factory=list)
    interfaces_summary: str = ""
    description: str = ""


def _resolve_proxy_display(raw_host: dict, proxy_lookup: Dict[str, str]) -> tuple[str, str]:
    monitored_by = str(raw_host.get("monitored_by", MONITORED_BY_SERVER))
    proxy_id = str(raw_host.get("proxyid", "0"))
    proxy_group_id = str(raw_host.get("proxy_groupid", "0"))

    if monitored_by == MONITORED_BY_PROXY and proxy_id != "0":
        name = proxy_lookup.get(proxy_id)
        return (name if name else f"Proxy (ID: {proxy_id})"), proxy_id
    if monitored_by == MONITORED_BY_PROXY_GROUP and proxy_group_id != "0":
        # Proxy groups (new in 7.0) are out of scope for name resolution in
        # this first version -- flagged clearly rather than mislabeled.
        return f"Proxy group (ID: {proxy_group_id})", ""
    return "Direct", ""


def _format_interface_line(iface: dict) -> str:
    itype = INTERFACE_TYPES.get(str(iface.get("type")), f"Unknown ({iface.get('type')})")
    ip = iface.get("ip", "") or ""
    dns = iface.get("dns", "") or ""
    port = iface.get("port", "") or ""
    is_main = str(iface.get("main")) == "1"

    location = ip if ip else ""
    if dns:
        location = f"{location} ({dns})" if location else dns
    if not location:
        location = "no address"

    line = f"{itype}: {location}"
    if port:
        line += f":{port}"
    if is_main:
        line += " [Main]"
    return line


def process_hosts(
    raw_hosts: List[dict], proxy_lookup: Dict[str, str]
) -> tuple[List[HostRow], List[dict], List[dict], List[dict], List[dict]]:
    """
    Normalize raw host.get results into:
      - a list of HostRow (for the 'Hosts' sheet)
      - flat interface rows (for the 'Interfaces' sheet)
      - flat host-group rows (for the 'Host Groups' sheet)
      - flat template rows (for the 'Templates' sheet)
      - flat tag rows (for the 'Tags' sheet)
    """
    LOG.info("Processing interfaces...")
    LOG.info("Processing host groups...")
    LOG.info("Processing templates...")
    LOG.info("Processing tags...")

    host_rows: List[HostRow] = []
    interface_rows: List[dict] = []
    group_rows: List[dict] = []
    template_rows: List[dict] = []
    tag_rows: List[dict] = []

    for raw in raw_hosts:
        hostid = raw.get("hostid", "")
        host_tech = raw.get("host", "")
        visible_name = raw.get("name", "") or host_tech
        status = HOST_STATUS.get(str(raw.get("status")), f"Unknown ({raw.get('status')})")

        if not raw.get("host"):
            LOG.warning("Host %s has no technical name; skipping display fallback issues.", hostid)

        proxy_display, proxy_id = _resolve_proxy_display(raw, proxy_lookup)

        interfaces = raw.get("interfaces") or []
        interface_lines = []
        for iface in interfaces:
            interface_lines.append(_format_interface_line(iface))
            interface_rows.append(
                {
                    "hostid": hostid,
                    "host": host_tech,
                    "type": INTERFACE_TYPES.get(str(iface.get("type")), f"Unknown ({iface.get('type')})"),
                    "ip": iface.get("ip", "") or "",
                    "dns": iface.get("dns", "") or "",
                    "port": iface.get("port", "") or "",
                    "main": "Yes" if str(iface.get("main")) == "1" else "No",
                }
            )
        if not interfaces:
            LOG.warning("Host '%s' (%s) has no configured interfaces.", host_tech, hostid)

        groups = raw.get("hostgroups") or []
        group_names = [g.get("name", "") for g in groups]
        for g in groups:
            group_rows.append({"hostid": hostid, "host": host_tech, "group": g.get("name", "")})
        if not groups:
            LOG.warning("Host '%s' (%s) is not a member of any host group.", host_tech, hostid)

        templates = raw.get("parentTemplates") or []
        template_names = [t.get("name", "") for t in templates]
        for t in templates:
            template_rows.append({"hostid": hostid, "host": host_tech, "template": t.get("name", "")})

        tags = raw.get("tags") or []
        tag_summaries = []
        for t in tags:
            tag_name = t.get("tag", "")
            tag_value = t.get("value", "") or ""
            tag_summaries.append(f"{tag_name}={tag_value}" if tag_value else tag_name)
            tag_rows.append({"hostid": hostid, "host": host_tech, "tag": tag_name, "value": tag_value})

        host_rows.append(
            HostRow(
                hostid=hostid,
                host=host_tech,
                name=visible_name,
                status=status,
                proxy_display=proxy_display,
                proxy_id=proxy_id,
                groups=group_names,
                templates=template_names,
                tags=tag_summaries,
                interfaces_summary="\n".join(interface_lines),
                description=raw.get("description", "") or "",
            )
        )

    host_rows.sort(key=lambda h: (h.host or "").lower())

    return host_rows, interface_rows, group_rows, template_rows, tag_rows


# ---------------------------------------------------------------------------
# 5. Excel generation
# ---------------------------------------------------------------------------

def _style_header(ws: Worksheet, num_cols: int) -> None:
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}1"


def _autosize_columns(ws: Worksheet, headers: List[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in row:
                if cell.value is None:
                    continue
                for line in str(cell.value).split("\n"):
                    max_len = max(max_len, len(line))
        width = min(max(max_len + 2, MIN_COL_WIDTH), MAX_COL_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _adjust_row_heights(ws: Worksheet, wrap_columns: List[int]) -> None:
    for row in ws.iter_rows(min_row=2):
        max_lines = 1
        for col_idx in wrap_columns:
            value = row[col_idx - 1].value
            if value:
                max_lines = max(max_lines, str(value).count("\n") + 1)
        if max_lines > 1:
            ws.row_dimensions[row[0].row].height = 15 * max_lines


def _write_sheet(wb: Workbook, title: str, headers: List[str], rows: List[List[Any]], wrap_columns: List[int]) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
        for col_idx in wrap_columns:
            ws.cell(row=ws.max_row, column=col_idx).alignment = WRAP_ALIGN_TOP
        for col_idx in range(1, len(headers) + 1):
            if col_idx not in wrap_columns:
                ws.cell(row=ws.max_row, column=col_idx).alignment = TOP_ALIGN

    _style_header(ws, len(headers))
    _autosize_columns(ws, headers)
    _adjust_row_heights(ws, wrap_columns)


def create_excel(
    host_rows: List[HostRow],
    interface_rows: List[dict],
    group_rows: List[dict],
    template_rows: List[dict],
    tag_rows: List[dict],
    output_path: str,
) -> None:
    LOG.info("Generating Excel file...")
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    hosts_headers = [
        "Host ID", "Host", "Visible Name", "Status", "Proxy",
        "Host Groups", "Templates", "Tags", "Interfaces", "Description",
    ]
    hosts_data = [
        [
            h.hostid,
            h.host,
            h.name,
            h.status,
            h.proxy_display,
            "\n".join(h.groups),
            "\n".join(h.templates),
            "\n".join(h.tags),
            h.interfaces_summary,
            h.description,
        ]
        for h in host_rows
    ]
    _write_sheet(wb, "Hosts", hosts_headers, hosts_data, wrap_columns=[6, 7, 8, 9])

    iface_headers = ["Host ID", "Host", "Interface Type", "IP", "DNS", "Port", "Main"]
    iface_data = [
        [r["hostid"], r["host"], r["type"], r["ip"], r["dns"], r["port"], r["main"]]
        for r in interface_rows
    ]
    _write_sheet(wb, "Interfaces", iface_headers, iface_data, wrap_columns=[])

    group_headers = ["Host ID", "Host", "Group"]
    group_data = [[r["hostid"], r["host"], r["group"]] for r in group_rows]
    _write_sheet(wb, "Host Groups", group_headers, group_data, wrap_columns=[])

    template_headers = ["Host ID", "Host", "Template"]
    template_data = [[r["hostid"], r["host"], r["template"]] for r in template_rows]
    _write_sheet(wb, "Templates", template_headers, template_data, wrap_columns=[])

    tag_headers = ["Host ID", "Host", "Tag", "Value"]
    tag_data = [[r["hostid"], r["host"], r["tag"], r["value"]] for r in tag_rows]
    _write_sheet(wb, "Tags", tag_headers, tag_data, wrap_columns=[])

    try:
        wb.save(output_path)
    except OSError as exc:
        raise ZabbixAPIError(f"Could not write Excel file to '{output_path}': {exc}") from exc

    LOG.info("Inventory successfully generated: %s", output_path)


# ---------------------------------------------------------------------------
# 6. Orchestration
# ---------------------------------------------------------------------------

def _configure_logging() -> None:
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s - %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )


def main(argv: Optional[List[str]] = None) -> int:
    _configure_logging()
    client: Optional[ZabbixClient] = None

    try:
        LOG.info("Loading configuration...")
        config = load_config(argv)

        client = create_zabbix_client(config)

        raw_hosts = get_hosts(client)
        proxy_lookup = get_proxy_lookup(client)

        host_rows, interface_rows, group_rows, template_rows, tag_rows = process_hosts(
            raw_hosts, proxy_lookup
        )

        create_excel(host_rows, interface_rows, group_rows, template_rows, tag_rows, config.output)

    except ConfigError as exc:
        LOG.error("Configuration error: %s", exc)
        return 2
    except ZabbixAPIError as exc:
        LOG.error("Zabbix API error: %s", exc)
        return 3
    except Exception as exc:  # noqa: BLE001 - top-level safety net, logged with context
        LOG.error("Unexpected error: %s", exc, exc_info=True)
        return 1
    finally:
        if client is not None:
            client.logout()

    return 0


if __name__ == "__main__":
    sys.exit(main())
